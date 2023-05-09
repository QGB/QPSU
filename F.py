# coding=utf-8
import sys as _sys;from os import path as _p#endswith 是为了适配qgb处于另外一个包内的情况
if __name__.endswith('qgb.F'):from . import py
else:import py
T=py.importT()
import os as _os
try:
	from io import BytesIO
	from io import BytesIO as bytesIO
	from io import BytesIO as bio
	from io import BytesIO as BIO
	from pathlib import Path
except:pass

gError=[]
def setErr(ae):
	global gError
	U=py.importU()
	if U.gbLogErr:# U.
		if type(gError) is list:gError.append(ae)
		elif gError:gError=[gError,ae]
		else:gError=[ae]
	else:
		gError=ae
	if U.gbPrintErr:U.pln('#Error ',ae)
	
def init_module():
	global DEFAULT_ENCODING		
	if DEFAULT_ENCODING is None:
		DEFAULT_ENCODING=U.get_or_set('F.read.encoding',lazy_default=lambda:py.No('utf-8'))
		
###################
class IntSize(py.int):
	def __new__(cls, *a, **ka):
	#int() argument must be a string, a bytes-like object or a number, not 
		
		if py.istr(a[0]) or py.isbyte(a[0]) or py.isnumber(a[0]):
			self= py.int.__new__(cls, *a)
		else:
			self= a[0]
		self.ka=ka
		return self
	def __str__(self):
		U,T,N,F=py.importUTNF()	
		# s=
		
		# repr=U.get_duplicated_kargs(self.ka,'repr','str','s','st','__repr__','__str__',no_pop=True)
		# if repr:
			# if py.callable(repr):
				# return repr(self, **self.ka )
			# else:
				# return py.str(repr)
			
		# return T.justify(s,**self.ka)+'>'
		
		str_size=U.get_duplicated_kargs(self.ka,'size','str_size','repr_size','s',default=0)
		return '<{}>'.format(int_to_size_str(self,str_size=str_size) )
		# return 
		# return '<{}={}>'.format(super().__repr__(),F.ssize(self) )
	def __repr__(self):return self.__str__()
################################
def gzip_decode(b):
	import zlib
	import urllib

	# f=urllib.request.urlopen(url) 
	decompressed_data=zlib.decompress(b, 16+zlib.MAX_WBITS)
	return decompressed_data	
decode_gzip_bytes=gzip_decode

def read_levelDB(db_dir,debug=0):
	'''
pip install plyvel-win32

conda install leveldb plyvel #不能用  ## plyvel.DB(db_dir) #进程退出 ！！
pip install plyvel 装不上

'''
	import plyvel
	U,T,N,F=py.importUTNF()	
	db_dir=F.auto_path(db_dir)
	db = U.get_or_set(db_dir,lazy_default=lambda:plyvel.DB(db_dir)) #只能打开一次 不然 IOError: b'IO error: 
	if db.closed:
		db=U.set(db_dir,plyvel.DB(db_dir))
	if debug:return db
	r={}
	with db.iterator() as it:
		for k, v in it:
			r[k]=v
			# pass	
	return r
read_leveldb=read_levelDB	

def get_file_owner_username(filename):
	''' Windows : ModuleNotFoundError: No module named 'pwd'
	'''
	import pwd,os
	return pwd.getpwuid(os.stat(filename).st_uid).pw_name
	
file_owner=file_username=get_file_user=get_file_username=get_file_owner_username
	
def zip(filename,fs,ext='.zip'):
	import zipfile,stat,os
	if not filename.lower().endswith(ext):
		filename+=ext

	with zipfile.ZipFile(filename, 'w') as zipMe:
		for file in fs:
			if os.path.islink(file):
				zipInfo  = zipfile.ZipInfo(file)
				zipInfo.create_system = 3 # System which created ZIP archive, 3 = Unix; 0 = Windows
				unix_st_mode = stat.S_IFLNK | stat.S_IRUSR | stat.S_IWUSR | stat.S_IXUSR | stat.S_IRGRP | stat.S_IWGRP | stat.S_IXGRP | stat.S_IROTH | stat.S_IWOTH | stat.S_IXOTH
				zipInfo.external_attr = unix_st_mode << 16 
				zipMe.writestr(zipInfo,os.readlink(file))
			else:
				zipMe.write(file, compress_type=zipfile.ZIP_DEFLATED)
	return zipMe
	
def compress_directory(source,target=py.No('auto use source name save in U.gst'),format='zip'):
	''' shutil.make_archive file: NotADirectoryError: [WinError 267] 目录名称无效。: './U.py'
shutil.make_archive(
    base_name,
    format,
    root_dir=None,
    base_dir=None,
    verbose=0,
    dry_run=0,
    owner=None,
    group=None,
    logger=None,
)
'''
	import shutil
	U,T,N,F=py.importUTNF()
	if not source.endswith('/'):source+='/'
	
	if not target:
		target=T.sub_last(F.get_dirname_from_full_path(source),'/','/')
		if not target:
			target=T.file_legalized(source)
	target=auto_path(target)
	if target.lower().endswith('.zip') and target[-5]!='/':
		target=target[:-4]
		
	return shutil.make_archive(target, format, source) #return target
zip_dir=zip_directory=compress_directory
	
def open(file,mode='r',**ka):
	'''py.open(
    file,
    mode='r',
    buffering=-1,
    encoding=None,
    errors=None,
    newline=None,
    closefd=True,
    opener=None,
)
'''
	U,T,N,F=py.importUTNF()
	mode=U.get_duplicated_kargs(ka,'mode','mod','m',default=mode)
	if py.isfile(file):
		#if file.closed==False:
		return file
	elif py.istr(file):
		return py.open(file,mode=mode,**ka)
	else:
		raise py.ArgumentUnsupported(file,ka)
	
def test_long_filename():
	'''
245
246
247
248
249 #FileNotFoundError     Windows 10                    

'''
	F=py.importF()
	v='a234567890b234567890c234567890d234567890e234567890f234567890g234567890h234567890i234567890j234567890k234567890l234567890m234567890n234567890o234567890p234567890q234567890r234567890s234567890t234567890u234567890v234567890w234567890x234567890y234567890z23456789'# len(v)==259
	p=F.mkdir(U.gst+'fn')
	U.cd(p)
	for n in range(245,260):
		print(n)
		with open(v[:n],'w') as f:f.write(str(n))
		with open(p+v[:n],'w') as f:f.write(str(n))
		#249 两个同样  #FileNotFoundError 
		
	# for n in range(245,260):
		# print(n)
def sub_head(f,a,b=b''):
	# br=
	if not py.isbyte(a):
		a=a.encode('utf-8')
	U,T,N,F=py.importUTNF()
	return F.write(f,T.sub(F.read_bytes(f),a,b))
sub=sub_head			
		
def replace(f,ba,bb,encoding='utf-8'):
	F=py.importF()
	if not py.isbytes(ba):ba=ba.encode(encoding)
	if not py.isbytes(bb):bb=bb.encode(encoding)
	
	br= F.read_bytes(f).replace(ba,bb)
	return F.write(f,br)
	
def expandUser(file='',user=''):
	'''always return endswith / 
	
import os, pwd
pwd.getpwuid(os.getuid()).pw_dir	
	'''
	import os
	h=os.path.expanduser('~'+user)
	h=auto_path(h)
	# if not r.endswith('/'):r=r+'/'
	T=py.importT()
	return T.replace_once(file,'~',h)
expand_user=expanduser=expandUser
	
def getHomeFromEnv():
	import os
	U=py.importU()
	if U.isLinux():
		name='HOME'
	elif U.isWin():
		name='USERPROFILE'
	else:
		raise EnvironmentError('#TODO system case')
	r=U.getEnv(name)
	r=autoPath(r)
	if not r.endswith('/'):r=r+'/'
	return r
home=gethome=get_home=getHome=getHomeFromEnv	
	
def include(file,keyword):
	if py.isbyte(keyword):mod='rb'
	else:mod='r'
	try:
		with py.open(file,mod) as f:
			for i in f:
				if keyword in i:return True
	except Exception as e:
		return py.No(e)
	return False
	
def stat(path, dir_fd=None, follow_symlinks=True):
	U=py.importU()
	IntSize,FloatTime,IntOct=U.IntSize,U.FloatTime,U.IntOct	
	import os
	try:
		if isinstance(path,os.stat_result):
			s=path
		else:
			s=os.stat(path=path, dir_fd=dir_fd, follow_symlinks=follow_symlinks)
		# return [
		# 		path,
		# 		IntSize(s.st_size),
		# 		FloatTime(s.st_atime),
		# 		FloatTime(s.st_mtime),
		# 		FloatTime(s.st_ctime),
		# 		IntOct (s.st_mode ),
		# 		]
		
	except Exception as e:
		return py.No(e)
	r={}
	for i in py.dir(s):
		if not i.startswith('st_'):continue
		v=getattr(s,i,py.No('Error getattr') )
		if i=='st_size':r[i]=IntSize(v);continue
		if i=='st_mode':r[i]=IntOct(v)      ;continue
		if i.endswith('time'):
			r[i]=FloatTime(v)
			continue
		r[i]=v
	return r
	
def repr_dill_load(obj):
	'''	不用考虑文件问题，以后要统一 序列化系列函数,
	但是这个函数序列化后是字符串，不能 autoArgs 是否是file参数
'''
	
def repr_dill_dump(obj):
	'''#TODO not dill basic type :  number,bytes,str,  ... (all ast.literal_eval) 
	'''
	
	return
	
def basic_dump(obj):
	return py.repr(obj)

def basic_load(sobj):
	T=py.importT()
	return T.unrepr(sobj)
	
def deSerialize(obj=None,file=None):
	'''The protocol version of the pickle is detected automatically, so no
protocol argument is needed.  Bytes past the pickled object's
representation are ignored.
'''
	if not py.isbyte(obj) and not file:raise py.ArgumentError('need bytes or file=str ')
	import pickle
	if py.istr(obj):
		file=obj
		obj=None
		U.log('autoArgs file=%s'%file)
	if py.isbyte(obj):
		return pickle.loads(obj)
	else:
		file=autoPath(file)
		with py.open(file,'rb') as f:
			return pickle.load(f)
ds=pickle_load=unSerialize=unserialize=deserialize=deSerialize
			
def serialize(obj,file=None,protocol=0):
	'''if not file: Return the pickled representation of the object as a bytes object.
	'''
	F=py.importF()
	try:
		import dill	
		return F.dill_dump(obj,file=file)
	except Exception as e:
		pass
	import pickle
	if file:
		file=autoPath(file)
		with py.open(file,'wb') as f:
			pickle.dump(obj=obj,file=f,protocol=protocol)
		return file
	else:
		return pickle.dumps(obj=obj,protocol=protocol)	
s=obj_dump=dump=pickle_dump=serialize

def dill_load_file(file,dill_ext='.dill'):
	import dill
	dill.settings['ignore']=False #KeyError: 'ignore'
	
	file=auto_path(file,ext=dill_ext)
	try:
		with py.open(file,'rb') as f:
			return dill.load(f)
	except Exception as e:#TODO all  load save py.No
		return py.No(file,e)
dl=dill_read=read_dill=dill_load=dill_load_file


def dill_load_bytes(bytes):
	import dill
	return dill.loads(bytes)
dls=dill_loads=dill_load_byte=dill_load_bytes

def dill_dump_bytes(obj,file=None,protocol=None,dill_ext='.dill'):
	'''
#TODO file=0  Not write '../0.dill'	
	
	dill.dump(obj, file, protocol=None, byref=None, fmode=None, recurse=None)
	dill.dumps(obj, protocol=None, byref=None, fmode=None, recurse=None)

	ValueError: pickle protocol must be <= 4 
r=request.get ...
F.readableSize(len(F.dill_dump(protocol=None,obj=r)  ) )#'14.192 KiB'

F.readableSize(len(F.dill_dump(protocol=0,obj=r)  ) )   #'15.773 KiB'
F.readableSize(len(F.dill_dump(protocol=1,obj=r)  ) )   #'19.177 KiB'
F.readableSize(len(F.dill_dump(protocol=2,obj=r)  ) )   #'18.972 KiB'
F.readableSize(len(F.dill_dump(protocol=3,obj=r)  ) )   #'14.192 KiB'
F.readableSize(len(F.dill_dump(protocol=4,obj=r)  ) )   #'13.694 KiB'

dill还包括几个pickle错误检测工具，在dill.detect module.
['at', 'baditems', 'badobjects', 'badtypes', 'children', 'code', 'dis', 'errors', 'freevars', 'getmodule', 'globalvars', 'iscode', 'isframe', 'isfunction', 'ismethod', 'istraceback', 'nestedcode', 'nestedglobals', 'outermost', 'parent', 'parents', 'reference', 'referredglobals', 'referrednested', 'trace', 'varnames', ...]
	'''
	import dill
	if file:
		if py.istr(obj) and py.len(obj)<333 and '.dill' in obj:
			if not py.istr(file) or '.dill' not in file:
				file,obj=obj,file
		file=auto_path(file,ext=dill_ext)
		with py.open(file,'wb') as f:
			dill.dump(obj=obj,file=f,protocol=protocol)		
		return file
	else:
		return dill.dumps(obj=obj,protocol=protocol)
dp=dumps=dill_write=write_dill=dill_dump=dill_dumps=dill_dump_bytes 

def dill_dump_string(obj,**ka):
	U=py.importU()
	encoding=U.get_duplicated_kargs(ka,'encoding','encode','coding')
	if not encoding:
		encoding=U.get_or_set('dill_string.encoding',default='latin')
	return dill_dump_bytes(obj).decode(encoding)	
dill_dump_str=dill_dump_string

def dill_load_string(s,**ka):
	U=py.importU()
	encoding=U.get_duplicated_kargs(ka,'encoding','encode','coding')
	if not encoding:
		encoding=U.get_or_set('dill_string.encoding',default='latin')
	return dill_load_bytes(s.encode(encoding) )
dill_load_str=dill_load_string

TRY_MAX_LAYER=5
def try_dill_dump_recursively(obj,*a,):
	global U
	if not U:U=py.importU()
	if py.len(a)>TRY_MAX_LAYER:return 
	try:
		b=dill_dump_bytes(obj)
		return (*a,U.size(b))
	except Exception as e:
		if py.islist(obj) or py.istuple(obj) or py.isset(obj):
			r=[]
			for n,v in py.enumerate(obj):
				r.append([n,try_dill_dump_recursively(v,*a,n)])
			return r
		elif py.isdict(obj):
			d={}
			for n,(k,v) in py.enumerate( obj.items()):
				d[n]=try_dill_dump_recursively(kv,*a,n)
			return d
		r=[]	
		for n,k,v in U.dir(obj):
			r.append([k,try_dill_dump_recursively(v,*a,k)])
		
tryDillDumpRecursively=recursive_try_dill_dump=try_dill_dump_recursively

def test_dir_recursively(obj,*a):
	U=py.importU()
	if py.len(a)>TRY_MAX_LAYER:return 
	r=U.dir(obj)
	for n,k,v in r:
		r[n][2]=test_dir_recursively(v,*a,k) or v
	return r
recursive_test_dir=test_dir_recursively

def recursive_test_dp(r):
	U=py.importU()
	if  py.isdict(r) or py.getattr(r,'items',0):
		r=[[n,k,v] for n,(k,v) in py.enumerate(r.items()) ]

	if py.islist(r) or py.isdict(r):
		pass
	else:return r
	
	for n,k,v in r:
		try:
			b=dill_dump_bytes(v)
			r[n][2]=U.size(b)
		except Exception as e:
			r[n][2]=recursive_test_dp(v)
	return r
	
def load(file,):
	''' '''
	
def write(file,obj,):
	''' '''
	
		
def chmod777(file,mode=0o777,):
	import os
	os.chmod(file, mode) 
chmod=chmod777

def getMode(file):
	import os
	try:
		r= oct(os.stat(file).st_mode)
		if r[:5]!='0o100':raise Exception('不是100代表什么？',r)
		return r[-3:]
	except Exception as e:
		return py.No(e)
getmode=	getMode
	
def copy_with_src_dir_struct(abs_src_dir,abs_dst_dir,symlinks=False, ignore=None):
	import shutil
	U,T,N,F=py.importUTNF()
	if U.isWin():raise NotImplementedError()
	
	if abs_src_dir[-1] != '/':abs_src_dir+='/'
	if not F.exist(abs_src_dir):return F.exist(abs_src_dir)
	
	if abs_dst_dir[-1] != '/':abs_dst_dir+='/'
	if not F.exist(abs_dst_dir):return F.exist(abs_dst_dir)
	
	if not abs_dst_dir.endswith(abs_src_dir):
		# if abs_src_dir.startswith('')
		abs_dst_dir+=abs_src_dir
		abs_dst_dir=abs_dst_dir.replace('//','/')
	return shutil.copytree(abs_src_dir,abs_dst_dir, symlinks=symlinks,ignore=ignore,)
copy_src_dir_struct=copy_with_src_dir_struct
	
def copy(src,dst,src_base='',skip=''):
	r''' src : sFilePath , list ,or \n strs
dst:sPath

return skip_list, copyed_list
	'''
	from shutil import copy as _copy
	U,T,N,F=py.importUTNF()
	
	if py.istr(skip):skip=[skip]
	if not py.istr(dst):raise py.ArgumentError('dst must be str')
	if py.istr(src):
		if '\n' in src:
			src=src.splitlines()
			return copy(src,dst)
		if src[-1] in ['/','\\']:
			src=F.ls(src,r=1)
		else:
			dst_dir=F.get_dirname_from_full_path(dst)
			F.mkdir(dst_dir)
			try:
				return _copy(src,dst)
			except Exception as e:
				return py.No(e)
	if not src_base:
		dl=U.unique(U.len(*src),ct=1)
		min=py.min(*dl)
		f=[i for i in src if py.len(i)==min][0]
		if f[-1] in ['/','\\']:f=f[:-1]
		# Path(f).absolute().parent.absolute().__str__()
		src_base=f[:py.len(T.sub_last(f.replace('\\','/'),'','/') )+1]
		src_base_len=py.len(src_base)
	print('src_base: %r'%src_base,'len(src)==%s'%py.len(src))
	while dst[-1] not in ['/','\\']:
		dst=U.input('not dir! rewrite dst:',default=dst)
	if py.iterable(src):
		fns=[]
		skips=[]
		for i in src:
			if U.one_in(skip,i):
				skips.append(i)
				continue
			# fn=getName(i)
			# if fn in fns:
				# fn=T.fileName(i)
			fn=i[src_base_len:]
			if fn[-1] in ['/','\\']:
				mkdir(dst+fn)
			else:	
				_copy(i,dst+fn)
			fns.append(fn)
		if skips:return skips,fns	
		return fns
	raise py.ArgumentUnsupported(src)
cp=copy
	
def modPathInSys(mod=None):
	if mod:
		if not py.istr(mod):mod=mod.__file__
	else:mod=__file__
	
	mod=mod.replace('\\','/')
	inPath=False
	if os.path.isabs(mod):
		for i in sys.path:
			i=i.replace('\\','/')
			if i and mod.startswith(i):
				if not i.endswith('/'):i+='/'
				i+='qgb'
				if mod.startswith(i):inPath=True
	else:
		raise NotImplementedError('__file__ not abs')
	return inPath
def lineCount(a):
	def blocks(files, size=65536):
		while True:
			b = files.read(size)
			if not b: break
			yield b

	with py.open(a, "r") as f:
		return sum(bl.count("\n") for bl in blocks(f))
		
def getPath(asp):
	asp=asp.replace('\\','/')
	if asp.endswith('/'):return asp
	else:
		if isDir(asp):return asp+'/'
		else:
			# _p.dirname('.../qgb/')# 'G:/QGB/babun/cygwin/lib/python2.7/qgb'
			# _p.dirname('.../qgb')# 'G:/QGB/babun/cygwin/lib/python2.7'
			return _p.dirname(asp)+'/'
def getPaths(a):
	r''' a:str 
	'''
	U=py.importU()
	if U.iswin():
		sp=a.replace('\\','/').split('/')
		rlist=[]
		def append(r):
			if r and r not in rlist:rlist.append(r)
		r=''
		for i,v in enumerate(sp):
			if len(v)>1 and v[-2] in T.AZ and v[-1]==':':
				append(r)
				r=v[-2:]+'/'
				continue

			for j in v:
				if j not in T.PATH_NAME:
					append(r)
					r=''
					continue

			if r and v:
				if isdir(r+v):r=r+v+'/'
				else:
					append(r)
					r=''
					continue
		return rlist
	else:
		raise NotImplementedError('*nux')
		
def get_filename_from_full_path(a):
	''' if a.endswith('/'):return ''
	
def name(	
	'''
	a=a.replace('\\','/')
	if '/' not in a:return a
	else:
		# import T
		return T.subr(a,'/','')
fileName=filename=get_name=getname=getName=getNameFromPath=getFilename=get_filename=get_filename_from_full_path
# filename=fileName=getname=getName=name
		
def getNameWithoutExt(a):
	''' see getNameFromPath
	'''
	
	a=getNameFromPath(a)
	if '.' in a:
		return T.subr(a,'','.')
	else:return a

def auto_find_file(head,ext='',r='Default auto accroding to the head'):
	'''r :recursion
	return str  
	# TODO # ext=?ext*     '''
	if not py.type(ext)==py.type(head)==py.str or head=='':
		return ''
	
	if len(ext)>0 and not ext.startswith('.'):ext='.'+ext
	head=head.lower();ext=ext.lower()
	head=head.replace('\\','/')
	if '/' in head:
		if py.type(r) is py.str:
			r=True
		else:
			r=False
	
	ap='.'
	if _p.isabs(head):
		ap=dir(head)
		if not isExist(ap):
			if head.endswith(ext):return head
			else:return head+ext
	
	# import F
	ls=[i.lower() for i in list(ap,r=r)]
	if head+ext in ls:return head+ext
	# import U
	# U.pprint(ls)
	for i in ls:		
		if i.startswith(head) and i.endswith(ext):return i
	
	for i in ls:
		if head in i and ext in i:return i
		
	# if inMuti(ext,'*?'):ext=ext.replace('*')
	
	if not head.endswith(ext):head+=ext
	
	return head
autof=auto_find_file
		
def new(a):
	'''will overwrite'''
	try:
		f=py.open(a,'w')
		f.write('')
		f.close()
		return f.name
	except Exception as e:
		setErr(e)
		return False
		
def isDir(ast):
	'''#TODO:
	
	'''	
	if not py.istr(ast):ast=py.str(ast)
	if not ast:return False
	ast=ast.replace('\\','/')
	if exist(ast):
		return _p.isdir(ast)
	else:
		if ast.endswith('/'):return True
		else:                return False
isPath=isdir=isDir
	# if not ast.replace('.','').strip():return True # （不是 点和空格 返回 T） 
	# return ('/' in ast) or ('\\' in ast)
	# return _p.sep in ast
# def is	
def bin(a,split=''):
	'''
bin(number, /)
F.bin(1.0)=='0b00111111100000000000000000000000'  # （大端）
'''
	import struct
	if py.isint(a):
		return py.bin(a)
	r='0b'
	if py.isfloat(a):
		r=r+split.join(py.bin(i).replace('0b', '').rjust(8, '0') for i in struct.pack('!f', a))
	elif py.isbytes(a):
		r=r+split.join(py.bin(i).replace('0b', '').rjust(8, '0') for i in a)
	else:
		raise py.ArgumentUnsupported('#TODO type',a)
	return r
def int_to_bytes(a):
	T=py.importU().T
	a=T.intToStr(a)
	return hexToBytes(a)
i2b=intToBytes=int_to_byte=int_to_bytes

def byte_to_int(a):
	return py.ord(a)
b2i=byte_to_int	

def bytes_to_hex(a,split=''):
	'''如果 len(split)是奇数，肯定返回 （奇数+偶数）=奇数
						偶=》偶
	'''
	if py.is3():
		# ord=lambda b:py.int.from_bytes(b,'big')
		ord=lambda i:i
	else:
		ord=py.ord
	r=split.join( [DIH[ord(i)] for i in a] )
	return r
b2h=bytesToHex=bytes_to_hex

def string_to_hex(a,encoding='utf-8',split=' '):
	if py.is_bytes(a):
		return bytes_to_hex(a,split=split)
	else:	
		return bytes_to_hex(a.encode(encoding),split=split)
s2h=str_hex=str_to_hex=string_to_hex	
	
def hexToBytes(a,split='',ignoreNonHex=True):
	a=a.upper();r=b''
	if ignoreNonHex:a=''.join([i for i in a if i in '0123456789ABCDEF'])
	it=2
	if len(split)>0:it+=len(split)
	if it==2 and len(a) % it!=0:return ()
	if it>2:
		a=a.split(split)
		if len(a[-1]) == 0:a=a[:-1]
		for i in a:
			r+=py.byte(DHI[ i ])
		return r
		
	for i in range(py.int(py.len(a)/2 )):
		r+=py.byte(DHI[a[i*2:i*2+2]])
	return r
h2b=hexToBytes


def writeIterable(file,data,end='\n',overwrite=True,encoding=None):
	U=py.importU()
	if not encoding:encoding=U.encoding
	
	file=autoPath(file)
	if overwrite:new(file)
	
	if py.is2():f=py.open(file,'a')
	else:       f=py.open(file,'a',encoding=encoding)
	
	for i in data:
		f.write(py.str(i)+end)
	f.close()
	return f.name	
	
def write(file,data,mod='w',encoding='utf-8',mkdir=False,autoArgs=True,pretty=True,seek=None):
	'''py3  open(file, mode='r', buffering=-1, encoding=None, errors=None, newline=None, closefd=True, opener=None)
	   py2  open(name[, mode[, buffering]])
pretty=True        Format a Python object into a pretty-printed representation.
	'''
	U=py.importU()
	try:
		if autoArgs:
			if py.istr(data) and py.len(file)>py.len(data)>0:
				if '.' in data and '.' not in file   and  isFileName(data):
					file,data=data,file
					U.warring('F.write fn,data but seems data,fn auto corrected（v 纠正')
	except:pass
	
	# try:
	file=autoPath(file)
	if not encoding:encoding=U.encoding
	
	if mkdir:makeDirs(file,isFile=True)
	
	# if 'b' not in mod and py.isbytes(data):mod+='b'# 自动检测 data与 mod 是否匹配
	
	if 'b' not in mod: #强制以 byte 写入
		mod+='b'
	f=py.open(file,mod)
		#f.write(强制unicode) 本来只适用 py.is3() ，但 py2 中 有 from io import open

	if py.isint(seek):
		f.seek(seek)
	# with open(file,mod) as f:	
	if py.isbyte(data):#istr(data) or (py.is3() and py.isinstance(data,py.bytes) )	:
		f.write(data)
	elif (py.is2() and py.isinstance(data,py.unicode)) :
		f.write(data.encode(encoding))
	elif (py.is3() and py.istr(data)):	
		# if 'b' in mod.lower():
		f.write(data.encode(encoding))
		# else:f.write(data)#*** UnicodeEncodeError: 'gbk' codec can't encode character '\xa9' in
	else:
		# if py.is2():print >>f,data
		# else:
		if pretty:
			data=U.pformat(data)
		U.pln(data,file=f)
	f.close()
	return f.name
	# except Exception as e:
		# setErr(e)
			# return False
			
gb_write_auto_filename_len=True
def write_auto_filename(*a,**ka):
	all_args=py.importU().getArgsDict()
	# py.pdb()()
	# return all_args
	U=py.importU()
	T=py.importT()
	name=U.get_duplicated_kargs(ka,'name',default=None)
	ext=U.get_duplicated_kargs(ka,'extension','ext',default='.txt')
	if ext and not ext.startswith('.'):ext='.'+ext
	
	sp=mkdir(U.gst+write_auto_filename.__name__)
	rf=[]
	for k,v in  all_args.items():
		if py.istuple(v) and py.len(v)==1:
			v=v[0]
	
		fn='{}{}'.format(sp,T.filename_legalized(k))
		if gb_write_auto_filename_len:
			len=U.len(v)
			if py.isint(len):
				fn+='={}{}'.format(len,ext)
		f=write(fn ,v,autoArgs=False)
		rf.append(f)
	return rf
writeA=write_auto_args=write_args=write_auto_filename

def insert_head_line(file,data,char_index=0):
	''' 没有很好的办法，除了 先全部读出再写入
	'''
	raise NotImplementedError()
    # f = fileinput.input(file, inplace=1)
    # for xline in f:
		# if f.isfirstline():
			# print line_to_prepend.rstrip('\r\n') + '\n' + xline,
		# else:
			# print xline,
	return
line_pre_adder=insert=insert_head_line
	
	
def append(file,data):
	'''builtin afile.write() No breakLine'''
	return write(file,data,mod='a')
	
def detect_file_encoding(file,confidence=0.7,default=py.No('not have default encoding'),buffer_size=9999,p=True,**ka):
	U,T,N,F=py.importUTNF()
	p=U.get_duplicated_kargs(ka,'print_file_encoding','print_detect_encoding','print',default=p,no_pop=True)
	if py.istr(file):
		with py.open(file,'rb') as f:
			b=f.read(buffer_size)
	elif py.isfile(file):
		if 'b' not in file.mode:raise py.ArgumentError("'b' not in file.mode",file)
		i=file.tell()
		b=file.read(buffer_size)
		file.seek(i)
		
	else:raise py.ArgumentError('need str or file')

	c= T.detect(b,confidence=confidence,default=default)
	if p:print(file,c) #TODO U.get_or_set('detect_file_encoding.p',True)
	return c
detect=detectEncoding=detect_encoding=detect_file_encoding
	
DEFAULT_ENCODING=None

def read(file,encoding='utf-8',mod='r',return_filename=False,print_detect_encoding=False,**ka):
	'''if return_filename:
			return content,f.name
1 not is 2
	   ^
SyntaxError: invalid syntax			
			'''

	file=autoPath(file)
	if not encoding and print_detect_encoding:
		U=py.importU()
		_pde=U.get_duplicated_kargs(ka,'print_encoding','p_encoding','p','pde','pEncoding','p_decode')
		if not _pde is U.GET_DUPLICATED_KARGS_DEFAULT: #记住绝对不能用 ==
			# print_detect_encoding=_pde
			print_detect_encoding=False
	if not return_filename:
		U=py.importU()
		return_filename=U.get_duplicated_kargs(ka,'returnFile','rf','rfn','return_file','return_name',)
		
	if py.is2():
		f=py.open(file,mod)
		s=f.read()
		f.close()
	else:#is3
		#utf-8 /site-packages/astropy/coordinates/builtin_frames/__init__.py  {'confidence': 0.73, 'encoding': 'Windows-1252'
		if encoding:
			try:
				f=py.open(file,mod,encoding=encoding)
				s=f.read()
				f.close()
			except:encoding=''
		if not encoding:
			U,T,N,F=py.importUTNF()	
			r2=T.detect_and_decode(F.read_byte(file),confidence=0.9,default='utf-8',return_encoding=True)
			if not r2:return r2
			U.set('r2',r2)
			encoding,s=r2
			if print_detect_encoding:print(file,encoding)
			
			
	
	if return_filename:
		return s,f.name
	else:
		return s
	# except Exception as e:
		# return f,e
		# if 'f' in py.dir() and f:f.close()
		# return ()
def read_multi_files_return_bytes_list(*fs,max_size=8*1024*1024,return_all_bytes=False):
	r=[]
	def append(a):	
		if return_all_bytes and not a:
			return r.append(b'')
		r.append(a)
		
	for f in fs:
		s=size_single_file(f)
		if not s :
			append(s)
			continue
		if s>max_size:
			append(py.No('f > max_size',f,max_size))
			continue
		with py.open(f,'rb') as fp:
			b=fp.read(max_size)
			append(b)
	return r		
read_multi_files=read_multi_file=read_multi_files_return_bytes=read_multi_files_return_bytes_list		
		
def read_bytes(file,size=-1,):
	'''is2 rb return str
f.read(size=-1, /)
Read and return up to n bytes.
	
	'''
	import io
	if isinstance(file, io.BytesIO):
		file.seek(0)
		return file.read(-1)
	file=autoPath(file)
	try:
		with py.open(file,'rb') as f:
			return f.read(size)
	except Exception as e:
		return py.No(e,file)
readb=readByte=readBytes=read_byte=read_bytes	


def read_bytes_chunks(path,start=0,chunk_size = 8192):
	'''try this func
except : StopIteration((PermissionError(13, 'Permission denied'), 'D:/',8192)
	'''
	if py.isfile(path):
		fd=path
		try:
			fd.read(0) #让错误提前暴露。
			fd.seek(start)
			while 1:
				buf = fd.read(chunk_size)
				if buf:
					yield buf
				else:
					break	
			return			
		except Exception as e:
			return (e,path,chunk_size)  
	path=auto_file_path(path)
	try:
		with open(path, 'rb') as fd:
			yield fd.read(0) # 让错误提前暴露。不会造成 500 Internal Server error
			fd.seek(start)
			while 1:
				buf = fd.read(chunk_size)
				if buf:
					yield buf
				else:
					break
	except Exception as e:
		# py.importU().print_tb_stack(path,chunk_size,e)
		return (e,path,chunk_size)          #               
		# return py.No(e,path,chunk_size)                         
		# raise StopIteration(e)
rbc=read_bytes_stream=read_file_as_stream=read_as_stream=read_file_chunks=readBytesChunks=read_bytes_chunks
			
def read_json(file,encoding=None):
	''' '''
	import json
	file=autoPath(file)
	if not encoding:encoding=detectEncoding(file)
	with py.open(file,encoding=encoding) as f:
		return json.load(f)
readjson=readJSON=json_load=read_json

def write_json(file,obj):
	import json
	file=auto_path(file,ext='json')
	with py.open(file,'w') as f: #not bytes,json write str
		json.dump(obj=obj,fp=f)
	return file

writeJSON=json_dump=write_json

def write_csv(file,rows,title=None):
	# import csv
	# with open(file, 'w') as f: 
		# write = csv.writer(f) 
		# if title:
			# write.writerow(title)
		# write.writerows(rows) 
		
	with open(file, 'w') as f:
		for row in rows:
			f.write('"{}","{}"\n'.format(*row))
	return file
def read_csv(file,encoding=None,delimiter=',',keep_default_na=False,):
	''' keep_default_na : use float nan ,not empty string ''
the  na_filter=False can change your columns type to object
	
	'''
	file=autoPath(file)
	if not encoding:encoding=detectEncoding(file)
	import pandas as pd
	df = pd.read_csv(file, delimiter=delimiter,encoding=encoding,keep_default_na=keep_default_na)
	r=[]
	is1=False
	if py.len(df.columns)==1:is1=True
	for i in df.values:
		if is1:
			r.append(i[0])
		else:
			r.append(tuple(i))
	return r
	# or export it as a list of dicts
	# dicts = df.to_dict().values()
readCSV=read_csv

def read_qpsu_file(file,prefix='file/'):
	U=py.importU()
	# 'E:/QGB/babun/cygwin/bin/qgb/'
	return read(U.getModPath()+prefix+file)
qpsufile=qpsuFile=readqp=readqpsu=readQPSU=read_qpsu=read_qpsu_file

def open_xlsx(file):
	from openpyxl import load_workbook
	wb = load_workbook(file)
	return wb

def write_xlsx(file,a):
	import openpyxl
	file=autoPath(file)
	if file.lower().endswith('.xls'):file=file[:-4]+'.xlsx'
	if not file.lower().endswith('.xlsx'):file=file+'.xlsx'
	
	outwb = openpyxl.Workbook()  # 打开一个将写的文件
	sheet = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
	for i, row in py.enumerate(a):
		for j, col in py.enumerate(row):
			sheet.cell(row=i+1, column=j+1).value=col
			# i,j=i+1,j+1 #我知道原因了，这是因为i+1 只应该每行执行一次，把它提到col循环外就行了
			# try:
				# sheet.cell(row=i, column=j).value=col #这样出现错误表格格式 ，原因？
			# except Exception as e:
				# return e,sheet,i,j,col
	outwb.save(file)  # 一定要记得保存
	return file
	
def read_xlsx(file,sheet=0):
	''' 加载一个8M多xlsx文件，很慢,'''
	import openpyxl
	if py.isinstance(file,openpyxl.workbook.workbook.Workbook):
		wb=file
	elif py.istr(file):
		wb = openpyxl.load_workbook(file)# 这句很耗时
	else:
		raise py.ArgumentUnsupported(file)
		
	if py.isint(sheet):
		ws=wb[wb.sheetnames[sheet]]
	elif py.istr(sheet):
		ws=wb[sheet]
	else:raise py.ArgumentError('sheet is index or sheet_names')	
	r=[]
	for row in ws.values:#<Worksheet "1 lemmas">
		ri=[]
		for value in row:
			ri.append(value)
		r.append(ri)	
	return r
		
def read_xlsx_sheets_name(file):
	''' 加载一个8M多xlsx文件，很慢,用了32秒
	'''
	from openpyxl import load_workbook
	wb = load_workbook(file)# 这句很耗时
	return wb.sheetnames
		
def write_xls(file,a):
	'''ValueError: row index was 65536, not allowed by .xls format'''
	import xlwt
	file=autoPath(file,ext='xls')
	
	xldoc = xlwt.Workbook()
	sheet = xldoc.add_sheet("Sheet1", cell_overwrite_ok=True)
	for i, row in py.enumerate(a):
		for j, col in py.enumerate(row):
			sheet.write(i, j, col)
	xldoc.save(file)
	return file

def read_xls(file,sheetIndex=0):
	''' return [ [colValue...]  .. ]  #No type description
	'''
	import xlrd                         
	w=xlrd.open_workbook(file)           
	sh=w.sheets()[sheetIndex]
	return sh._cell_values
		
def read_xls_sheets_name(file):
	''' XLRDError: Excel xlsx file; not supported'''
	import xlrd                         
	w=xlrd.open_workbook(file)           
	return py.list(py.enumerate( w.sheet_names() )  )
get_xls_sheets_name=read_xls_sheets_name
	
	
SQL_DEFAULT="SELECT * FROM {};"	
def read_sqlite(file,table='',sql=SQL_DEFAULT,limit=None):
	file=autoPath(file)
	if table:sql=sql.format(table)
	import sqlite3
	with sqlite3.connect(file) as con:
		cursor = con.cursor()

		if table or sql!=SQL_DEFAULT:
			# if not (table in tables):return py.No('no table',table,'found in',file)
			cursor.execute(sql )
			return cursor.fetchall()
		else:		
			cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
			#[('sessions',), ('sqlite_sequence',), ('history',), ('output_history',)]
			tables=[i[0] for i in cursor.fetchall()]
			r={}
			for i in tables:
				r[i]=read_sqlite(file=file,sql=SQL_DEFAULT.format(i))
			return r
readSqlite=read_sqlite

def readYaml(file):
	import yaml
	with py.open(file) as f:
		return yaml.safe_load(f)

def writeYaml(file,obj):
	import yaml
	try:
		with py.open(file,'w') as f:
			yaml.dump(obj,f,default_flow_style=False)#default_flow_style=False parameter is necessary to produce the format you want (flow style), otherwise for nested collections it produces block style:
			return file
	except Exception as e:
		return py.No(e,file,obj)
		
def isDir(file):
	return _p.isdir(file)
is_folder=isFolder=is_dir=isdir=isDir
	
def exists(fn,zero=True):
	''': path.exist('c:')
Out[57]: False

In [58]: path.exist('o:')#这是因为当前目录下存在 'o:'，但是只能用ls完整查看
Out[58]: True
## 在cygwin中 只能判断 c:\ 或 c:/ ,单独 c: 总为False
#############
>>> path.exists('c:')
True
>>> path.exists('o:')
True
'''
	if not fn:return fn
	U=py.importU()
	if U.iscyg():
		if len(fn)==2 and fn[-1]==':':fn+='/'
		# raise NotImplementedError
	if _p.exists(fn):
		if _p.isdir(fn):
			# fn=fn.replace('\\','/')
			if not (fn.endswith('/') or  fn.endswith('\\') ):fn+=get_splitor(fn)
			return fn
		if _p.getsize(fn)<1 and not zero:
			return zero
		return fn
	else:
		return py.No(fn,'not exists')
isExist=exist=exists

def glob(path='./', pattern='**/*'):
	'''pattern='**/*'   : all files
	'**/*.txt'   : all txt files'''
	import pathlib
	return py.list(pathlib.Path(path).glob(pattern))
wildcard_find=find_file_wildcard=glob
	
def walk(ap):
	'''
return tuple(ap,dir_list,file_list )
	2&3 : os.walk(top, topdown=True, onerror=None, followlinks=False)
	'''
	if py.is2():
		return _os.walk(ap).next()
	else:
		return _os.walk(ap).__next__()
		
def list(ap='.',type='',t='',r=False,d=False,dir=False,f=False,
	file=False,include='',exclude='',timeout=None,print_result=False,**ka):
	'''Parms:bool r recursion
			 str (type,t) '(d,f,a,r)'
	default return all'''
	U=py.importU()
	print_result=U.get_duplicated_kargs(ka,'print_r','print','p',default=print_result)
	
	if dir:d=True
	if file:f=True
	if t and not type:type=t
	
	if 'd' in type:d=True
	if 'f' in type:f=True
	if 'a' in type:d=True;f=True
	if 'r' in type:r=True
	
	if d or dir or f or file:pass
	else:d=f=True		#default return all
	
	if not py.istr(ap) or py.len(ap)<1:
		setErr('F.list arguments ap error')
		ap='.'
	# if len(ap)==2 and ap.endswith(':'):ap+='/'	
	if U.isnix() and ap.startswith('~'):
		if ap[1:2]=='/':ap=ap[2:]
		else:ap=ap[1:]
		ap=get_home()+ap
	ap=ap.replace('\\','/')
	if not ap.endswith('/'):#U.inMuti(ap,'/','\\',f=str.endswith):
		if isDir(ap):
			ap+='/'
		else:

			return [exists(ap)]
		# if not ap.endswith('/'):ap+='/'
		# else:ap+='\\'
	
	# U.repl()
	########## below r is result
	rls=[]
	try:r3=py.list(walk(ap))
	except Exception as ew:
		# pln ap;raise ew
		return py.No(ew)
	
	if ap=='./':ap=''
	# U.repl()
	r3[1]=[ap+i+'/' for i in r3[1]] #dirs
	r3[2]=[ap+i for i in r3[2]] #files
	
	
	if d:rls.extend(r3[1])
 
	# 
	if r:
		for i in r3[1]:rls.extend(list(i,r=r,d=d,f=f))
			
	if f:rls.extend(r3[2])
	
	if include:rls=[i for i in rls if include in i]
	if exclude:rls=[i for i in rls if exclude not in i]
	if print_result:
		U.pprint(rls)
	return rls
	# else:return r3[1]+r3[2]
ls=list

def get_filemode(f):
	U=py.importU()
	return U.IntOct(_os.stat(f).st_mode )
mod=mode=get_mod=get_mode=get_filemode	

def ll(ap='.',readable=True,type='',t='',r=False,d=False,dir=False,f=False,file=False,
	return_dict=True,return_list=False,no_raise=True,**ka):
	'''return {file : [size,atime,mtime,ctime,st_mode]}	
	readable is True: Size,Stime,..
	linux struct stat: http://pubs.opengroup.org/onlinepubs/009695399/basedefs/sys/stat.h.html'''
	U=py.importU()
	return_list=U.get_duplicated_kargs(ka,'return_list','list','rl','l',default=return_list)
	if return_list:return_dict=False
	
	dr={}
	
	for i in list(ap,type=type,t=t,r=r,d=d,dir=dir,f=f,file=file):#ls 肯定返回 list类型！
		# importU;if U.DEBUG:U.pln ap,repr(i)
		try:
			s=_os.stat(i)
		except:
			continue
		if readable:
			
			IntSize,FloatTime,IntOct=U.IntSize,U.FloatTime,U.IntOct
			try:
				dr[i]=[ IntSize(size(i)),
						FloatTime(s.st_atime),
						FloatTime(s.st_mtime),
						FloatTime(s.st_ctime),
						IntOct (s.st_mode ),
					]
			except Exception as e:
				if no_raise:
					dr[i]=py.No(e)
				else:
					raise
				
		else:
			dr[i]=[size(i),s.st_atime,s.st_mtime,s.st_ctime,s.st_mode]
	if return_dict or not return_list:
		return dr
	else:
		return [[k,*v] for k,v in dr.items()]

SUFFIXES = {1000: ['KB', 'MB', 'GB', 'TB', 'PB', 'EB', 'ZB', 'YB'],
			1024: ['KiB', 'MiB', 'GiB', 'TiB', 'PiB', 'EiB', 'ZiB', 'YiB']}
def int_to_size_str(size,b1024=True,zero='0 B',less_than_zero='%s',str_size=0,):
	'''Convert a file size to human-readable form.
	Keyword arguments:
	size -- file size in bytes
	b1024 -- if True (default), use multiples of 1024
	if False, use multiples of 1000
	Returns: string
	test git
	'''
	U,T,N,F=py.importUTNF()	
	if py.istr(size) or py.isbyte(size):
		size=U.len(size)
	size=py.int(size)
	def padding_str(a):
		if not str_size:return a
		sn,su=a.split(' ')
		i=str_size-py.len(sn)-py.len(su)
		i=py.max(i,0)
		return sn+' '*i+su
	if size < 0:
		if py.callable(less_than_zero):
			return less_than_zero(size)
		if py.istr(less_than_zero):
			if '%' in less_than_zero:
				return less_than_zero%size
			if '{' in less_than_zero and '}' in less_than_zero:
				return less_than_zero.format(size)
		return less_than_zero
	if size==0:return padding_str(zero)
		# raise ValueError('number must be non-negative')

	multiple = 1024.0 if b1024 else 1000.0 #another if statement style
	
	if size<=multiple:return padding_str(py.str(size)+' B'  )
	
	for suffix in SUFFIXES[multiple]:
		size /= multiple
		if size < multiple:
			return padding_str('{0:.3f} {1}'.format(size, suffix)  )
	raise ValueError('number too large')
rsize=ssize=readable_size=readablesize=readableSize=numToSize=int_to_size=int_to_size_str

def size_single_file(f,repr_size=0):
	f=nt_path(f)
	size =0 #0L  SyntaxError in 3
	if not _p.exists(f):
		return py.No('{} NOT EXISTS!'.format(f))
	if _p.isdir(f):
		# if f==''  , _p use pwd
		return py.No('{} is dir !'.format(f))
	
	try:
		size= _p.getsize(f)
		if size<=0:
			size=py.len(read_bytes(f))
		F=py.importF()	
		return F.IntSize(size,size=repr_size)
	except Exception as e:
		return py.No('unexcept err',e,f)
		
get_single_file_size=single_file_size=size_single_file
	
def size(asf,int=py.No('ipython auto readable'),repr_size=0): 
	'''file or path return byte count
	
repr_size=12  才能整齐
	'''
	asf=nt_path(asf)#if linux etc ,will auto ignored
	# asf=autoPath(asf)#in there,can't use gst
	size =0 #0L  SyntaxError in 3
	if not _p.exists(asf):
		return py.No('{} NOT EXISTS!'.format(asf))
	if not _p.isdir(asf):
		size= _p.getsize(asf)
		if size<=0:
			try:size=len(read_bytes(asf))
			except Exception as e:
				return py.No('unexcept err',e,asf)
		# return size
	else:# is dir
		for root, dirs, files in _os.walk(asf):  
			size += sum([_p.getsize(_p.join(root, name)) for name in files])  	
	U,T,N,F=py.importUTNF()
	# U.msgbox(U.is_ipy_cell())
	if not int :#and U.is_ipy_cell():
		size=F.IntSize(size,repr_size=repr_size)
	return size 
	
# U.pln size(U.gst)
# exit()
	
def csvAsList(fn):
	import csv
	with py.open(fn, 'rb') as f:
		reader = csv.reader(f)
		return py.list(reader)

def getSourcePath():
	f=_sys._getframe().f_back
	def back(af,ai=0):
		if af is None:return ''
		if '__file__' in af.f_globals:return af.f_globals['__file__']
		else:return back(af.f_back,ai+1)
		
	return back(f)
	
	# U.pln globals().keys()
	# U.pln __file__, __name__ , __package__
	
def delete_dir(a,raise_exception=False):
	''' shutil.rmtree 无法删除 只读文件 （例如 .git 目录下的）报 PermissionError(13, '拒绝访问。') e.filename
'''	
	import shutil,os,stat
	
	if raise_exception:
		shutil.rmtree(a)
		return a
		
	try:
		shutil.rmtree(a, )# ignore_errors=True  这里不发出 异常 
		return a	
	except py.PermissionError as e:
		file=e.filename
		if not os.access(file, os.W_OK):
			os.chmod(file, stat.S_IWUSR)
		########	
		try:
			return 	delete_dir(a,raise_exception=True)			
		except Exception as e:
			return py.No(e)
		########
	except Exception as e:
		return py.No(e)
		
def deleteFile(file):
	file=autoPath(file)
	sp=getSplitor(file)
	# for i in a.split(ap):
	U=py.importU()
	def Error():
		#Error()  返回None时   
		#TypeError: catching classes that do not inherit from BaseException is not allowed
		return FileNotFoundError
		#issubclass(FileNotFoundError,WindowsError) == True
		if U.iswin():	return WindowsError 
	
	try:
		if U.isWin():
			return py.from_qgb_import('Win').shell_delete(file)
	
		if isDir(file):
			return delete_dir(file)
		_os.remove(file)#异常 是从这里产生的
		return file
	except FileNotFoundError as e:
		if isDir(file):
			raise Exception('#TODO')
		return py.No(file,e,'Not exists?')
	#Error  {'G:\\test\\npp': WindowsError(5, '')}
	# Docstring: MS-Windows OS system call failed.  
	#*nix NameError: name 'WindowsError' is not defined  '''
	except Exception as e:
		# setErr({file:e})
		return py.No(file,e)
	return False
rm=delete=deleteFile
	
def getSplitor(ap):
	if '/' in ap:return '/'
	if '\\' in ap:return '\\'		
	return '/'#default	
getsp=getS=get_splitor=getSplitor		
		
def join_path(a,*p):
	'''os.path.join(a, *p)
Docstring:
Join two or more pathname components, inserting '/' as needed.
If any component is an absolute path, all previous path components
will be discarded.  An empty last part will result in a path that
ends with a separator.
File:      /root/anaconda3/lib/python3.7/posixpath.py
'''
	import os
	return os.path.join(a, *p)
join=joinPath=path_join=join_path

def mdcd(ap):
	return py.importU().cd(makeDirs(ap))
def move_back(target,source,**ka):
	return move(source,target,**ka)
def move(source,target,edit_target=False,mkdir=True,remove_invalid_char=True,**ka):
	''' * in target == source
? in target == 	source_fn
// in target == path(exclude source_fn)

# os.rename(source, target) #OSError: [WinError 17] 系统无法将文件移到不同的磁盘驱动器。: 'C:
 os.rename target文件名（包不包括路径都没关系）包括扩展名，最大不能超过241

	'''
	U,T,N,F=py.importUTNF()
	edit_target=U.get_duplicated_kargs(ka,'e','edit','editarget',default=edit_target)
	remove_invalid_char=U.get_duplicated_kargs(ka,'remove_char','delChar',
'del_char','remove_illegal','del__illegal','del_invalid','remove_invalid',
'del_invalid_char','del__illegal_char','remove_illegal_char',default=remove_invalid_char)
	source=F.autoPath(source)
	source_fn=F.get_filename_from_full_path(source)
	source_path=source[:-py.len(source_fn)]
	
	if target.endswith('/'):
		target+=source_fn
	# if '*' in target:
	target=target.replace('*',source)
	target=target.replace('?',source_fn)
	target=target.replace('//',source_path)
	
	target=F.autoPath(target)	
	
	if edit_target:target=U.input(default=target)
	if remove_invalid_char:
		if U.is_windows():
			target=T.replacey(target,T.NOT_PATH_NAME_WINDOWS,'')
	if not target.endswith('/') and F.isDir(target):
		target=target+'/'
	if mkdir:F.mkdir(F.get_dir(target)) # 要先有文件夹，不然shutil.move有可能找不到
	import shutil
	try:
		return shutil.move(source, target) 
		# return target
	except (FileNotFoundError,FileExistsError) as e: # parentheses required, otherwise invalid syntax
		return py.No(e,source,target)
		
	if F.isDir(target):
		r=target+source_fn
	else:
		r=target
	
	return F.exist(r)	
	
	# if not F.exist(target):
	# 	return py.No('moved to '+ target+' ,But not exist',source)
	# import shutil
	# shutil.move(source, target)
	# os.replace(source, target)
mv=move

def makeDirs(ap,isFile=False,cd=0,no_auto_path=False):
	''' 访问移动硬盘时，可能出现已经创建成功，但是 F.ls 看不到的情况。
	用explorer访问后又正常
	
	'''
	U=py.importU()
	if not no_auto_path:ap=autoPath(ap)
	if not py.isbool(isFile) and not py.isint(isFile):
		U.log('F.md(str,isFile={})'.format(repr(isFile)))
	if py.is3():
		from pathlib import Path
		p=Path(ap).absolute()
		
		if isFile:
			return makeDirs(p.parent,isFile=False)
		if p.is_file():#if not exists, is_dir() is_file() both return False
			return py.No(ap+' exists , and it is a file')
		r=py.No('unexpected err')
		try:
			p.mkdir()
		except (FileNotFoundError,) as e:
			if p.parent==p:# 'D:\\' 驱动器不存在
				r=e
			else:
				r=makeDirs(p.parent,isFile=False)
				if r:p.mkdir() # 建立父目录后，再次尝试创建本目录 
				else:return r
			# else:return r
			#但是如果父目录本来是个文件，再mkdir则FileNotFoundError: [WinError 3] 系统找不到指定的路径。
		except FileExistsError:
				
			pass
		except Exception as e:
			r=e
		if p.exists():
			if not no_auto_path:sp=autoPath(p)
			else:sp=py.str(p.absolute())
			splitor=get_splitor(sp)
			# else:sp=py.str(p.absolute()).replace('\\','/')
			
			if p.is_dir() and not sp.endswith(splitor):
				sp+=splitor
			r=sp
		else:
			r=py.No(r,p)
		if r and cd:U.cd(r)
		return r

	# if py.is2():		#########################################	
	U=py.importU()
	sp=getSplitor(ap)
	ap=ap.split(sp)
	if isFile:ap=ap[:-1]
	# if not isabs(ap):
		# if not ap.startswith('.'):
			# if ap.startswith(sp):ap=U.gst[:-1]+ap
			# else:ap=U.gst+ap
	# else:
	base=''
	
	for i in ap:
		base+=(i+sp)
		#TODO 2019-1-19  未处理存在的是文件  而不是 文件夹的情况
		if exist(base):continue
		else:
			try:
				_os.mkdir(base)
			
			except Exception as e:
				if U.iswin():
					if e.winerror==5:continue#WindowsError(5, '') 拒绝访问
					if e.winerror==183:continue#WindowsError 183 当文件已存在时，无法创建该文件。
				if 'exists' in e.args[1]:#(17, 'File exists') cygwin;
					continue
				# U.repl(printCount=True)
				setErr(e)
				return False
	return True	
	
	# if U.iswin():
		
		# _os.system('md "{0}"'.format(ap))
md=mkdir=makeDirs		

gbAutoPath=True
def auto_path(fn,ext='',default='',is_dir=False,p=False):
	''' default is U.gst 
if fn.startswith("."): 如果路径中所有文件夹存在，则可以写入读取。否则无法写入读取。file io 都是这个规则吧
FileNotFoundError: [Errno 2] No such file or directory: '.

#TODO #BUG# F.auto_path('\\\\\\C:\\test\\clipboard',default='C:/test/',)== '///C:/test/clipboard'   

	'''
	U=py.importU()
	def auto_string():
		nonlocal fn,ext
		fn=fn.replace('\\','/')
		if ext and not ext.startswith('.'):ext='.'+ext
		if not fn.lower().endswith(ext.lower()):fn+=ext
		if fn.startswith("~/"):
			import os
			if U.isnix():
				home=os.getenv('HOME')
			else:
				home=os.getenv('USERPROFILE')# 'C:/Users/Administrator'  not  cyg home os.getenv('HOME')
			# else:		home=os.getenv('HOMEPATH')#  HOMEPATH=\Users\Administrator
			fn= home+fn[1:];

		#TODO to avoid dos max_path ,  must \ full path 2019年12月19日 不记得什么意思？ 
		if (not fn.startswith(".")) and (not isAbs(fn)) :
			while fn.startswith('/'):fn=fn[1:]
			fn= default + fn
		
		if py.len(fn)>=260 and U.iswin():
			fn=nt_path(fn)
		if(is_dir and not fn.endswith('/')):fn+='/'
		# return fn
	
	if not gbAutoPath:
		pass
	else:
		if default:
			default=default.replace('\\','/')
			if not default.endswith('/'):default+='/'
		else:
			default=U.set_test_path(U.gst) # 防止 U.gst 被改变没被保存
		fn=str(fn)
		if py.istr(fn) and (fn.startswith(r'\\192.') or fn.startswith('//192.168') or fn.startswith(r'\\?\UNC')):
			pass#todo more 
		else:
			# fn=auto_string(fn)
			auto_string()
	
	if p:print(fn)
	return fn
autofn=auto_filename=autoFileName=auto_file_path=autoPath=auto_path


def get_nt_short_path_name(long_name,max=250):
	'''
win32api.GetShortPathName 一定要用 '\\'分隔符，不能用 '/' ，否则：
error: (2, 'GetShortPathNameW', '系统找不到指定的文件。')

#TODO SMB IP network file
'''	
	import win32api
	if py.len(long_name)<max:return long_name
	long_name=long_name.replace('/','\\')
	if not long_name.startswith( u"\\\\?\\"):
		long_name=u"\\\\?\\"+long_name
	return win32api.GetShortPathName(long_name)
shortPath=short_path=get_win_short_path=win_short_path=get_windows_short_path=windows_short_path=nt_short_path=get_short_path=GetShortPath=GetShortPathName=get_nt_short_path_name	

def nt_path(fn):
	'''if linux etc ,will auto ignored
write success, read failed ?	

 
	'''
	U=py.importU()
	fn=_p.abspath(fn)
	if  U.iswin():#or cyg?
		if not py.isunicode(fn):fn=fn.decode(U.T.detect(fn))#暂时没想到更好方法，头晕
		fn=fn.replace(u'/',u'\\')
		if u"\\\\?\\" in fn:
			pass
		elif fn.startswith(u"\\\\"):
			fn=u"\\\\?\\UNC\\" + fn[2:]
		else:
			fn=u"\\\\?\\" + fn    
	return U.StrRepr(fn)
	
def abs_path(file):
	'''In [165]: gs ###notice Users/Admin
Out[165]: '\\\\?\\C:\\Users/Administrator\\.gradle\\caches\\3.3\\scripts-remappe
d\\sync_local_repo10406_a5s4kku7mncoj5pzsbgck2y4v\\6oxfw7eb7mpz692y3xnywccj1\\in
it1efd45104ffa2d33563b85b9edda76e3\\classes\\sync_local_repo10406_a5s4kku7mncoj5
pzsbgck2y4v$_run_closure1$_closure2$_closure4$_closure5.class'

In [166]: F._p.abspath gs
--------> F._p.abspath(gs)
Out[166]: '\\\\?\\C:\\Users\\Administrator\\.gradle\\caches\\3.3\\scripts-remapp
ed\\sync_local_repo10406_a5s4kku7mncoj5pzsbgck2y4v\\6oxfw7eb7mpz692y3xnywccj1\\i
nit1efd45104ffa2d33563b85b9edda76e3\\classes\\sync_local_repo10406_a5s4kku7mncoj
5pzsbgck2y4v$_run_closure1$_closure2$_closure4$_closure5.class'''
	return _p.abspath(file).replace('\\','/')
abs=abspath=abs_path		

def isAbs(file):
	'''in cygwin:
In [43]: U.path.isabs( 'M:/Program Files/.babun/cygwin/lib/python2.7/qgb/file/attr.html')
Out[43]: False

In [45]: U.path	.isabs('M:\\Program Files\\.babun\\cygwin\\lib\\python2.7\\qgb\\file\\attr.html' )
Out[45]: False
##### win py3.6    fixed
--------> F.isabs('2d:\1\2')# False
--------> F.isabs('d:\1\2')# False
--------> F.isabs('/babun/cygwin/lib/python2.7\\qgb\\F.py')# True
#TODO 更加精细判断，文件名不合法 return None?
'''	
	U=py.importU()
	if U.iscyg() or U.iswin():
		if ':' in file:return True
		else:return False
	return _p.isabs(file)
isabs=isAbs
		
def name(a):
	'''Anti abs
	得到 单独文件（夹）名字
	'''
	U=py.importU()
	if not U.T.istr(a):return ''
	# if U.inMuti(a,'/','\\',f=str.endswith):a=a[:-1]
	if a.endswith('/') or a.endswith('\\'):a=a[:-1]
	if not isAbs(a):return a
	else:
		
		a=T.sub(a,dir(a))
		# U.repl()
		if a.endswith('/') or a.endswith('\\'):return a[1:]
		else:return a
		
def get_dirname_from_full_path(a):
	#TODO 判断 磁盘上a 为文件夹 应该直接返回
	if a[-1] in ['/','\\']:
		return a
	r= _p.dirname(a)
	if r in ['/','\\']:
		return r
	else:
		return r+get_splitor(r) 
	# exit()
get_dir=dirname=dir=get_path_from_full_path=get_dirname_from_full_path

def get_parent_dir(a):
	raise py.NotImplementedError()
	return
	
# def open(file, mode='r', buffering=-1, encoding=None, errors=None, newline=None, closefd=True, opener=None):
	'''py2 from io import open
----> 1 open(file, mode='r', buffering=-1, encoding=None, errors=None, newline=None, closefd=True, opener=None)

TypeError: open() takes at most 7 arguments (8 given)
	TypeError: 'opener' is an invalid keyword argument for this function
	
'''


def readlines(a,EOL=True,encoding=None,str_repr=False):
	a=autoPath(a)
	if not encoding:
		encoding=detectEncoding(a)
	def _return(lines):
		if str_repr:
			U=py.importU()
			return [U.StrRepr(i) for i in lines]
		return lines
	try:
		if EOL:
			r=[]
			for i in py.open(a,encoding=encoding):r.append(i)
		else:
			r=read(a,encoding=encoding).splitlines()
		return _return(r)
	except Exception as e:
		return py.No(e)
		
# F.isFileName('g:/a')
# []
def isFileName(a):
	'''a:str'''
	# from . import U #ValueError: Attempted relative import in non-package
	U=py.importU()
	for i in a:
		if i not in T.PATH_NAME:return py.No(i,a)
		
	return a
	

gdINT_HEX=DIH=DINT_HEX=INT_HEX={0:'00',1:'01',2:'02',3:'03',4:'04',5:'05',6:'06',7:'07',8:'08',9:'09',10:'0A',11:'0B',12:'0C',13:'0D',14:'0E',15:'0F',16:'10',17:'11',18:'12',19:'13',20:'14',21:'15',22:'16',23:'17',24:'18',25:'19',26:'1A',27:'1B',28:'1C',29:'1D',30:'1E',31:'1F',32:'20',33:'21',34:'22',35:'23',36:'24',37:'25',38:'26',39:'27',40:'28',41:'29',42:'2A',43:'2B',44:'2C',45:'2D',46:'2E',47:'2F',48:'30',49:'31',50:'32',51:'33',52:'34',53:'35',54:'36',55:'37',56:'38',57:'39',58:'3A',59:'3B',60:'3C',61:'3D',62:'3E',63:'3F',64:'40',65:'41',66:'42',67:'43',68:'44',69:'45',70:'46',71:'47',72:'48',73:'49',74:'4A',75:'4B',76:'4C',77:'4D',78:'4E',79:'4F',80:'50',81:'51',82:'52',83:'53',84:'54',85:'55',86:'56',87:'57',88:'58',89:'59',90:'5A',91:'5B',92:'5C',93:'5D',94:'5E',95:'5F',96:'60',97:'61',98:'62',99:'63',100:'64',101:'65',102:'66',103:'67',104:'68',105:'69',106:'6A',107:'6B',108:'6C',109:'6D',110:'6E',111:'6F',112:'70',113:'71',114:'72',115:'73',116:'74',117:'75',118:'76',119:'77',120:'78',121:'79',122:'7A',123:'7B',124:'7C',125:'7D',126:'7E',127:'7F',128:'80',129:'81',130:'82',131:'83',132:'84',133:'85',134:'86',135:'87',136:'88',137:'89',138:'8A',139:'8B',140:'8C',141:'8D',142:'8E',143:'8F',144:'90',145:'91',146:'92',147:'93',148:'94',149:'95',150:'96',151:'97',152:'98',153:'99',154:'9A',155:'9B',156:'9C',157:'9D',158:'9E',159:'9F',160:'A0',161:'A1',162:'A2',163:'A3',164:'A4',165:'A5',166:'A6',167:'A7',168:'A8',169:'A9',170:'AA',171:'AB',172:'AC',173:'AD',174:'AE',175:'AF',176:'B0',177:'B1',178:'B2',179:'B3',180:'B4',181:'B5',182:'B6',183:'B7',184:'B8',185:'B9',186:'BA',187:'BB',188:'BC',189:'BD',190:'BE',191:'BF',192:'C0',193:'C1',194:'C2',195:'C3',196:'C4',197:'C5',198:'C6',199:'C7',200:'C8',201:'C9',202:'CA',203:'CB',204:'CC',205:'CD',206:'CE',207:'CF',208:'D0',209:'D1',210:'D2',211:'D3',212:'D4',213:'D5',214:'D6',215:'D7',216:'D8',217:'D9',218:'DA',219:'DB',220:'DC',221:'DD',222:'DE',223:'DF',224:'E0',225:'E1',226:'E2',227:'E3',228:'E4',229:'E5',230:'E6',231:'E7',232:'E8',233:'E9',234:'EA',235:'EB',236:'EC',237:'ED',238:'EE',239:'EF',240:'F0',241:'F1',242:'F2',243:'F3',244:'F4',245:'F5',246:'F6',247:'F7',248:'F8',249:'F9',250:'FA',251:'FB',252:'FC',253:'FD',254:'FE',255:'FF'}
gdHEX_INT=DHI=DHEX_INT=HEX_INT={y:x for x,y in DIH.items()}
gdhb=Dhb=hex_byte=Dhex_bytes={'00':b'\x00','01':b'\x01','02':b'\x02','03':b'\x03','04':b'\x04','05':b'\x05','06':b'\x06','07':b'\x07','08':b'\x08','09':b'\x09','0a':b'\x0a','0b':b'\x0b','0c':b'\x0c','0d':b'\x0d','0e':b'\x0e','0f':b'\x0f','10':b'\x10','11':b'\x11','12':b'\x12','13':b'\x13','14':b'\x14','15':b'\x15','16':b'\x16','17':b'\x17','18':b'\x18','19':b'\x19','1a':b'\x1a','1b':b'\x1b','1c':b'\x1c','1d':b'\x1d','1e':b'\x1e','1f':b'\x1f','20':b'\x20','21':b'\x21','22':b'\x22','23':b'\x23','24':b'\x24','25':b'\x25','26':b'\x26','27':b'\x27','28':b'\x28','29':b'\x29','2a':b'\x2a','2b':b'\x2b','2c':b'\x2c','2d':b'\x2d','2e':b'\x2e','2f':b'\x2f','30':b'\x30','31':b'\x31','32':b'\x32','33':b'\x33','34':b'\x34','35':b'\x35','36':b'\x36','37':b'\x37','38':b'\x38','39':b'\x39','3a':b'\x3a','3b':b'\x3b','3c':b'\x3c','3d':b'\x3d','3e':b'\x3e','3f':b'\x3f','40':b'\x40','41':b'\x41','42':b'\x42','43':b'\x43','44':b'\x44','45':b'\x45','46':b'\x46','47':b'\x47','48':b'\x48','49':b'\x49','4a':b'\x4a','4b':b'\x4b','4c':b'\x4c','4d':b'\x4d','4e':b'\x4e','4f':b'\x4f','50':b'\x50','51':b'\x51','52':b'\x52','53':b'\x53','54':b'\x54','55':b'\x55','56':b'\x56','57':b'\x57','58':b'\x58','59':b'\x59','5a':b'\x5a','5b':b'\x5b','5c':b'\x5c','5d':b'\x5d','5e':b'\x5e','5f':b'\x5f','60':b'\x60','61':b'\x61','62':b'\x62','63':b'\x63','64':b'\x64','65':b'\x65','66':b'\x66','67':b'\x67','68':b'\x68','69':b'\x69','6a':b'\x6a','6b':b'\x6b','6c':b'\x6c','6d':b'\x6d','6e':b'\x6e','6f':b'\x6f','70':b'\x70','71':b'\x71','72':b'\x72','73':b'\x73','74':b'\x74','75':b'\x75','76':b'\x76','77':b'\x77','78':b'\x78','79':b'\x79','7a':b'\x7a','7b':b'\x7b','7c':b'\x7c','7d':b'\x7d','7e':b'\x7e','7f':b'\x7f','80':b'\x80','81':b'\x81','82':b'\x82','83':b'\x83','84':b'\x84','85':b'\x85','86':b'\x86','87':b'\x87','88':b'\x88','89':b'\x89','8a':b'\x8a','8b':b'\x8b','8c':b'\x8c','8d':b'\x8d','8e':b'\x8e','8f':b'\x8f','90':b'\x90','91':b'\x91','92':b'\x92','93':b'\x93','94':b'\x94','95':b'\x95','96':b'\x96','97':b'\x97','98':b'\x98','99':b'\x99','9a':b'\x9a','9b':b'\x9b','9c':b'\x9c','9d':b'\x9d','9e':b'\x9e','9f':b'\x9f','a0':b'\xa0','a1':b'\xa1','a2':b'\xa2','a3':b'\xa3','a4':b'\xa4','a5':b'\xa5','a6':b'\xa6','a7':b'\xa7','a8':b'\xa8','a9':b'\xa9','aa':b'\xaa','ab':b'\xab','ac':b'\xac','ad':b'\xad','ae':b'\xae','af':b'\xaf','b0':b'\xb0','b1':b'\xb1','b2':b'\xb2','b3':b'\xb3','b4':b'\xb4','b5':b'\xb5','b6':b'\xb6','b7':b'\xb7','b8':b'\xb8','b9':b'\xb9','ba':b'\xba','bb':b'\xbb','bc':b'\xbc','bd':b'\xbd','be':b'\xbe','bf':b'\xbf','c0':b'\xc0','c1':b'\xc1','c2':b'\xc2','c3':b'\xc3','c4':b'\xc4','c5':b'\xc5','c6':b'\xc6','c7':b'\xc7','c8':b'\xc8','c9':b'\xc9','ca':b'\xca','cb':b'\xcb','cc':b'\xcc','cd':b'\xcd','ce':b'\xce','cf':b'\xcf','d0':b'\xd0','d1':b'\xd1','d2':b'\xd2','d3':b'\xd3','d4':b'\xd4','d5':b'\xd5','d6':b'\xd6','d7':b'\xd7','d8':b'\xd8','d9':b'\xd9','da':b'\xda','db':b'\xdb','dc':b'\xdc','dd':b'\xdd','de':b'\xde','df':b'\xdf','e0':b'\xe0','e1':b'\xe1','e2':b'\xe2','e3':b'\xe3','e4':b'\xe4','e5':b'\xe5','e6':b'\xe6','e7':b'\xe7','e8':b'\xe8','e9':b'\xe9','ea':b'\xea','eb':b'\xeb','ec':b'\xec','ed':b'\xed','ee':b'\xee','ef':b'\xef','f0':b'\xf0','f1':b'\xf1','f2':b'\xf2','f3':b'\xf3','f4':b'\xf4','f5':b'\xf5','f6':b'\xf6','f7':b'\xf7','f8':b'\xf8','f9':b'\xf9','fa':b'\xfa','fb':b'\xfb','fc':b'\xfc','fd':b'\xfd','fe':b'\xfe','ff':b'\xff'}



if __name__=='__main__':
	import T,U
	# l=csvAsList('process.csv')
	# U.shtml(l)
	# U.pln getSourcePath()
	# U.repl()
	exit()	
	r=hexToBytes(T.HEX.lower())
	# U.write('h.r',r)
	